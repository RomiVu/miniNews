from io import BytesIO
from collections import defaultdict

from fastapi import APIRouter, File, Form, UploadFile, HTTPException, status, Depends
from openpyxl import load_workbook

from ..dependency import get_mysql
from ..db.crud import create_by_file


router = APIRouter()

@router.get("/{task_id}")
async def read_task(task_id: int):
    return {"id": task_id, "task_info": "http:www.baidu.com/"}


@router.post('/files')
async def upload_file(
    fileb: UploadFile = File(...),
    remark: str = Form(...),
    db=Depends(get_mysql)
):
    if fileb.filename.endswith('.xlsx') or fileb.filename.endswith('.csv'):
        file_like = await fileb.read()
        wb = load_workbook(filename=BytesIO(file_like))

        inserted = defaultdict(int)
        for ws in wb.sheetnames:
            _ws = wb[ws]
            if ws == "user":
                for i in range(2, _ws.max_row + 1):
                    inserted[_ws] += 1
                    username = _ws[f'a{i}'].value
                    password = _ws[f'b{i}'].value
                    await create_by_file(db, ws, {'username':username, 'password':password})
            elif ws == "tasks":
                for i in range(2, _ws.max_row + 1):
                    inserted[_ws] += 1
                    url = _ws[f'a{i}'].value
                    description = _ws[f'b{i}'].value
                    user_id = _ws[f'c{i}'].value
                    await create_by_file(db, ws, {'url':url, 'description':description, 'user_id': user_id})
            elif ws == "client":
                for i in range(2, _ws.max_row + 1):
                    inserted[_ws] += 1
                    organization = _ws[f'a{i}'].value
                    address = _ws[f'b{i}'].value
                    remark = _ws[f'c{i}'].value
                    await create_by_file(db, ws, {'organization':organization, 'address':address, 'remark': remark})
            else:
                pass

        db.commit()

        return {
            "filename": fileb.filename,
            "worksheet": wb.sheetnames,
            "file_size": len(file_like),
            "token": remark,
            "inserted": inserted.items(),
            "fileb_content_type": fileb.content_type,
        }
    raise HTTPException(
        status_code=status.HTTP_400_BAD_REQUEST,
        detail="only .xlsx and .csv allowed"
    )